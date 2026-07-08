#!/usr/bin/env python3
"""
=============================================================================
Solace Cloud / SAP AEM — User & Role Export Script
=============================================================================
Description : Exports all users and their assigned roles from a Solace Cloud
              or SAP AEM organisation via the Cloud REST API.
              Outputs: CSV + Excel (.xlsx) + JSON

API Used    : GET https://api.solace.cloud/api/v2/platform/users
Auth        : Bearer Token (generate: Cloud Console → User Icon → Token Management)
Pagination  : pageSize=100, iterates pages until nextPage is null

Usage:
    # Simplest — token via environment variable, everything else defaulted
    export SOLACE_API_TOKEN="your-token-here"
    python3 solace_cloud_export_script.py

    # Token via CLI flag
    python3 solace_cloud_export_script.py --token "your-token-here"

    # Settings via config.yaml (see config.example.yaml)
    cp config.example.yaml config.yaml   # edit with your token(s)
    python3 solace_cloud_export_script.py

    # Select a named profile from config.yaml (e.g. a second org / SAP AEM)
    python3 solace_cloud_export_script.py --profile sap_aem

Config precedence (highest wins): CLI flag > config.yaml (--profile section,
falling back to the top-level defaults) > SOLACE_API_TOKEN env var (token
only) > built-in default. config.yaml is entirely optional — the script
runs the same as before if it is absent.

Requirements:
    python3 -m pip install -r requirements.txt   # requests, pandas, openpyxl, PyYAML

Notes:
  - Confirmed working against Solace Cloud org "<tenant>" (US region, 2026-06-30)
  - Confirmed working unmodified against a second, SAP AEM org (2026-06-30) —
    see README section "SAP AEM — Live Cross-Platform Validation"
  - SAP AEM uses the IDENTICAL API base URL and endpoints — only the bearer
    token differs. Documented regional base URLs vary slightly between
    Solace sources; verify live before using a non-US region (see README).
  - /api/v2/iam/users returned 404; use /api/v2/platform/users (confirmed 200 OK)
  - Roles are embedded in each user object — no separate per-user role fetch needed
=============================================================================
"""

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import requests
except ImportError:
    sys.exit("❌ Missing: python3 -m pip install requests")

try:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("❌ Missing: python3 -m pip install pandas openpyxl")

try:
    import yaml
except ImportError:
    sys.exit("❌ Missing: python3 -m pip install PyYAML")


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION — built-in fallbacks; overridable via config.yaml or CLI args
# ─────────────────────────────────────────────────────────────────────────────

BASE_URL        = "https://api.solace.cloud"   # US region
USERS_ENDPOINT  = "/api/v2/platform/users"
PAGE_SIZE       = 100
TIMESTAMP       = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")   # yyyymmddhhMMss
VALID_FORMATS   = ("csv", "excel", "json", "all")


# ─────────────────────────────────────────────────────────────────────────────
# CONFIG FILE LOADING
# ─────────────────────────────────────────────────────────────────────────────

def load_config(path: str) -> dict:
    """
    Load config.yaml if present. config.yaml is entirely optional here —
    unlike a broker-connection tool, this script can run purely off CLI
    flags / the SOLACE_API_TOKEN env var. Returns {} when the file is
    missing so callers can fall through to other sources.
    """
    config_path = Path(path)
    if not config_path.exists():
        return {}
    with open(config_path) as f:
        return yaml.safe_load(f) or {}


def resolve_settings(args, config: dict) -> dict:
    """
    Merge CLI args, config.yaml (optionally scoped to --profile), and
    built-in defaults into a single settings dict.

    Precedence: CLI flag > config.yaml profile/root > env var (token only)
    > built-in default.
    """
    scope = dict(config)
    if args.profile:
        profile_cfg = (config.get("profiles") or {}).get(args.profile)
        if profile_cfg is None:
            sys.exit(f"❌ Profile '{args.profile}' not found in {args.config}")
        scope = {**config, **profile_cfg}

    token = args.token or scope.get("token") or os.environ.get("SOLACE_API_TOKEN")
    base_url = (args.base_url or scope.get("base_url") or BASE_URL).rstrip("/")
    output_dir = args.output_dir or scope.get("output_dir") or "./output"
    fmt = args.format or scope.get("format") or "all"
    role_sep = args.role_separator if args.role_separator is not None else scope.get("role_separator", " | ")

    if fmt not in VALID_FORMATS:
        sys.exit(f"❌ Invalid format '{fmt}' (config.yaml or --format). Choose from: {', '.join(VALID_FORMATS)}")

    return {
        "token": token,
        "base_url": base_url,
        "output_dir": output_dir,
        "format": fmt,
        "role_separator": role_sep,
    }


# ─────────────────────────────────────────────────────────────────────────────
# ARGUMENT PARSING
# ─────────────────────────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(
        description="Export Solace Cloud / SAP AEM users and roles"
    )
    parser.add_argument("--token", "-t",
        help="API Bearer token (overrides config.yaml / SOLACE_API_TOKEN env var)", default=None)
    parser.add_argument("--base-url", "-u",
        help=f"API base URL (overrides config.yaml; default: {BASE_URL})", default=None)
    parser.add_argument("--output-dir", "-o",
        help="Output directory (overrides config.yaml; default: ./output)", default=None)
    parser.add_argument("--format", "-f",
        help="Output format: csv | excel | json | all (overrides config.yaml; default: all)",
        choices=VALID_FORMATS, default=None)
    parser.add_argument("--role-separator",
        help="Separator for multiple roles in CSV (overrides config.yaml; default: ' | ')", default=None)
    parser.add_argument("--config", "-c",
        help="Path to config.yaml (default: config.yaml; file is optional)", default="config.yaml")
    parser.add_argument("--profile", "-p",
        help="Named profile from config.yaml's 'profiles' section (e.g. sap_aem)", default=None)
    return parser.parse_args()


# ─────────────────────────────────────────────────────────────────────────────
# API FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def get_headers(token: str) -> dict:
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type":  "application/json",
        "Accept":        "application/json",
    }


def fetch_all_users(base_url: str, token: str, role_sep: str = " | ") -> list[dict]:
    """
    Fetch all users from Solace Cloud API with full pagination.

    Pagination (v2 platform API):
      meta.pagination.nextPage  → next page number, or null (= last page)
      meta.pagination.count     → total user count in org
      meta.pagination.totalPages → total page count
    """
    url     = f"{base_url}{USERS_ENDPOINT}"
    headers = get_headers(token)
    all_users = []
    page = 1
    total_pages = None

    print(f"\n📡 Endpoint : {url}")
    print(f"   Page size : {PAGE_SIZE} users/page\n")

    while True:
        params = {"pageSize": PAGE_SIZE, "pageNumber": page}
        print(f"   ↳ Fetching page {page}" +
              (f" of {total_pages}" if total_pages else " (probing…)") + " …", end=" ")

        try:
            resp = requests.get(url, headers=headers, params=params, timeout=30)
            resp.raise_for_status()
        except requests.exceptions.HTTPError:
            print(f"\n❌ HTTP {resp.status_code}: {resp.text}")
            sys.exit(1)
        except requests.exceptions.ConnectionError:
            print(f"\n❌ Connection error — check BASE_URL: {base_url}")
            sys.exit(1)
        except requests.exceptions.Timeout:
            print(f"\n❌ Timeout on page {page}")
            sys.exit(1)

        payload    = resp.json()
        pagination = payload.get("meta", {}).get("pagination", {})
        total_pages = pagination.get("totalPages", "?")
        next_page   = pagination.get("nextPage")
        total_count = pagination.get("count", "?")

        users_this_page = payload.get("data", [])
        print(f"{len(users_this_page)} users")

        if page == 1:
            print(f"\n   ✅ Organisation total : {total_count} users across {total_pages} page(s)\n")

        for user in users_this_page:
            roles_list = user.get("roles", [])
            all_users.append({
                "user_id":       user.get("id", ""),
                "organization":  user.get("organizationId", ""),
                "first_name":    user.get("firstName", ""),
                "last_name":     user.get("lastName", ""),
                "email":         user.get("email", ""),
                "state":         user.get("state", ""),
                "roles":         role_sep.join(sorted(roles_list)),
                "role_count":    len(roles_list),
                "groups":        ", ".join(user.get("groups", [])),
                # Boolean helpers for quick filtering
                "is_admin":      "administrator" in roles_list,
                "is_billing_admin": "billing-administrator" in roles_list,
            })

        if next_page is None:
            break
        page = next_page

    # Sort by email (case-insensitive)
    all_users.sort(key=lambda u: u["email"].lower())
    print(f"\n✅ Total users retrieved : {len(all_users)}\n")
    return all_users


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def write_csv(users: list[dict], filepath: str):
    """Write user list to CSV."""
    if not users:
        print("⚠️  No users to write.")
        return
    df = pd.DataFrame(users)
    df.to_csv(filepath, index=False, encoding="utf-8-sig")  # utf-8-sig for Excel compat
    print(f"📄 CSV   → {filepath}  ({len(users)} rows)")


def write_excel(users: list[dict], filepath: str):
    """
    Write user list to Excel (.xlsx) with:
    - Frozen header row
    - Bold + coloured header
    - Auto-fitted column widths
    - Separate 'Admins' sheet for administrator users
    - Separate 'Role Summary' sheet with role frequency count
    """
    if not users:
        print("⚠️  No users to write.")
        return

    df = pd.DataFrame(users)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        # Sheet 1: All users
        df.to_excel(writer, sheet_name="All Users", index=False)
        _format_sheet(writer, "All Users", df)

        # Sheet 2: Admins only
        admins_df = df[df["is_admin"] == True].copy()
        admins_df.to_excel(writer, sheet_name="Admins", index=False)
        _format_sheet(writer, "Admins", admins_df)

        # Sheet 3: Role frequency summary
        role_counts = {}
        for user in users:
            for role in user["roles"].split(" | "):
                role = role.strip()
                if role:
                    role_counts[role] = role_counts.get(role, 0) + 1
        role_df = pd.DataFrame(
            sorted(role_counts.items(), key=lambda x: -x[1]),
            columns=["role", "user_count"]
        )
        role_df.to_excel(writer, sheet_name="Role Summary", index=False)
        _format_sheet(writer, "Role Summary", role_df)

    print(f"📊 Excel → {filepath}  ({len(users)} users, 3 sheets)")


def _format_sheet(writer, sheet_name: str, df: pd.DataFrame):
    """Apply formatting to an Excel worksheet."""
    ws = writer.sheets[sheet_name]
    header_fill = PatternFill(start_color="1F5C99", end_color="1F5C99", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:  # Row 1 = headers
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    # Auto-fit column widths (max 60 chars)
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)


def write_json(users: list[dict], filepath: str):
    """Write user list to JSON."""
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump({
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "total_users": len(users),
            "users": users
        }, f, indent=2, ensure_ascii=False)
    print(f"🗂️  JSON  → {filepath}  ({len(users)} records)")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    args = parse_args()
    config = load_config(args.config)
    settings = resolve_settings(args, config)

    if not settings["token"]:
        sys.exit(
            "❌ No API token provided.\n"
            "   Set env var:   export SOLACE_API_TOKEN='your-token'\n"
            "   Or use flag:   --token 'your-token'\n"
            "   Or configure:  cp config.example.yaml config.yaml   # then edit"
        )

    # Each run gets its own timestamped subdirectory under output_dir, e.g.
    # ./output/20260708142530/ — keeps successive exports from overwriting
    # or interleaving with each other.
    run_dir = os.path.join(settings["output_dir"], TIMESTAMP)
    os.makedirs(run_dir, exist_ok=True)

    print("=" * 60)
    print("  Solace Cloud / SAP AEM — User & Role Export")
    print("=" * 60)
    print(f"  Config file: {args.config}" + (" (not found — using CLI/env only)" if not config else ""))
    if args.profile:
        print(f"  Profile    : {args.profile}")
    print(f"  Base URL   : {settings['base_url']}")
    print(f"  Output dir : {os.path.abspath(run_dir)}")
    print(f"  Format(s)  : {settings['format']}")
    print(f"  Run at     : {datetime.now(timezone.utc).isoformat()}")
    print("=" * 60)

    # Fetch all users
    users = fetch_all_users(settings["base_url"], settings["token"], role_sep=settings["role_separator"])

    if not users:
        print("⚠️  No users returned. Check token permissions.")
        sys.exit(0)

    # Write outputs — filenames no longer need a timestamp suffix since the
    # parent directory already carries the run's timestamp.
    fmt = settings["format"]
    if fmt in ("csv", "all"):
        write_csv(users,  os.path.join(run_dir, "solace_users_roles.csv"))
    if fmt in ("excel", "all"):
        write_excel(users, os.path.join(run_dir, "solace_users_roles.xlsx"))
    if fmt in ("json", "all"):
        write_json(users,  os.path.join(run_dir, "solace_users_roles.json"))

    print("\n✅ Export complete!")
    print(f"   Files written to: {os.path.abspath(run_dir)}/")


if __name__ == "__main__":
    main()
