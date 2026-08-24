"""
Unit and integration tests for solace_cloud_export_script.py / verify_export.py.

The exporter has no --mock mode (unlike some other Solace tooling in this
family of repos) — it always calls a real, authenticated API. So instead of
hitting a live Solace Cloud org, these tests mock the HTTP layer
(requests.get) with small, fully synthetic fixture data. No network access
and no real credentials are needed to run them.

Run with:
    pip install -r requirements-dev.txt
    pytest
"""

import csv
import json
import sys
from argparse import Namespace
from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import solace_cloud_export_script as export_mod  # noqa: E402
import verify_export as verify_mod  # noqa: E402


# ---------------------------------------------------------------------------
# Fixture data — a small, fully synthetic two-page user list
# ---------------------------------------------------------------------------

PAGE_1_USERS = [
    {"id": "u1", "organizationId": "test-org", "firstName": "Alice", "lastName": "Anders",
     "email": "alice.anders@example.com", "roles": ["administrator"], "groups": [], "state": "ACTIVE"},
    {"id": "u2", "organizationId": "test-org", "firstName": "Bob", "lastName": "Brown",
     "email": "bob.brown@example.com", "roles": ["billing-administrator", "event-portal-user"],
     "groups": ["finance"], "state": "ACTIVE"},
]
PAGE_2_USERS = [
    {"id": "u3", "organizationId": "test-org", "firstName": "", "lastName": "",
     "email": "carol.chen@example.com", "roles": ["event-portal-user"], "groups": [], "state": "INVITED"},
]


def _make_response(users, page_number, total_pages, next_page):
    resp = MagicMock()
    resp.status_code = 200
    resp.raise_for_status = MagicMock()
    resp.json.return_value = {
        "data": users,
        "meta": {"pagination": {
            "pageNumber": page_number, "pageSize": 100,
            "count": len(PAGE_1_USERS) + len(PAGE_2_USERS),
            "totalPages": total_pages, "nextPage": next_page,
        }},
    }
    return resp


def _fake_get(url, headers=None, params=None, timeout=None):
    page = params["pageNumber"]
    if page == 1:
        return _make_response(PAGE_1_USERS, 1, 2, 2)
    if page == 2:
        return _make_response(PAGE_2_USERS, 2, 2, None)
    raise AssertionError(f"Unexpected page requested: {page}")


# ---------------------------------------------------------------------------
# resolve_settings() — precedence: CLI > config/profile > env var > default
# ---------------------------------------------------------------------------

def _args(**overrides):
    defaults = dict(token=None, base_url=None, output_dir=None, format=None,
                     role_separator=None, config="config.yaml", profile=None)
    defaults.update(overrides)
    return Namespace(**defaults)


def test_resolve_settings_builtin_defaults(monkeypatch):
    monkeypatch.delenv("SOLACE_API_TOKEN", raising=False)
    settings = export_mod.resolve_settings(_args(), {})
    assert settings["token"] is None
    assert settings["base_url"] == export_mod.BASE_URL
    assert settings["output_dir"] == "./output"
    assert settings["format"] == "all"
    assert settings["role_separator"] == " | "


def test_resolve_settings_env_var_used_when_nothing_higher_set(monkeypatch):
    monkeypatch.setenv("SOLACE_API_TOKEN", "env-token")
    settings = export_mod.resolve_settings(_args(), {})
    assert settings["token"] == "env-token"


def test_resolve_settings_config_wins_over_env_var(monkeypatch):
    monkeypatch.setenv("SOLACE_API_TOKEN", "env-token")
    settings = export_mod.resolve_settings(_args(), {"token": "config-token"})
    assert settings["token"] == "config-token"


def test_resolve_settings_cli_wins_over_config_and_env(monkeypatch):
    monkeypatch.setenv("SOLACE_API_TOKEN", "env-token")
    settings = export_mod.resolve_settings(_args(token="cli-token"), {"token": "config-token"})
    assert settings["token"] == "cli-token"


def test_resolve_settings_profile_overrides_top_level_defaults():
    config = {
        "token": "root-token", "output_dir": "./output",
        "profiles": {"sap_aem": {"token": "aem-token", "output_dir": "./output/sap-aem"}},
    }
    settings = export_mod.resolve_settings(_args(profile="sap_aem"), config)
    assert settings["token"] == "aem-token"
    assert settings["output_dir"] == "./output/sap-aem"


def test_resolve_settings_profile_falls_back_to_root_for_missing_keys():
    config = {
        "token": "root-token", "base_url": "https://api.solace.cloud",
        "profiles": {"sap_aem": {"token": "aem-token"}},  # no base_url override
    }
    settings = export_mod.resolve_settings(_args(profile="sap_aem"), config)
    assert settings["base_url"] == "https://api.solace.cloud"


def test_resolve_settings_unknown_profile_exits():
    with pytest.raises(SystemExit):
        export_mod.resolve_settings(_args(profile="does-not-exist"), {})


def test_resolve_settings_invalid_format_from_config_exits():
    # argparse's choices= blocks a bad --format at the CLI, but a bad value
    # in config.yaml reaches resolve_settings() directly — must be caught here.
    with pytest.raises(SystemExit):
        export_mod.resolve_settings(_args(), {"format": "yaml"})


# ---------------------------------------------------------------------------
# fetch_all_users() — pagination, flattening, admin flags, sorting
# ---------------------------------------------------------------------------

def test_fetch_all_users_paginates_and_sorts_by_email(monkeypatch):
    monkeypatch.setattr(export_mod.requests, "get", _fake_get)
    users = export_mod.fetch_all_users("https://api.solace.cloud", "fake-token")
    assert [u["email"] for u in users] == [
        "alice.anders@example.com", "bob.brown@example.com", "carol.chen@example.com",
    ]


def test_fetch_all_users_sets_admin_flags(monkeypatch):
    monkeypatch.setattr(export_mod.requests, "get", _fake_get)
    users = export_mod.fetch_all_users("https://api.solace.cloud", "fake-token")
    by_email = {u["email"]: u for u in users}
    assert by_email["alice.anders@example.com"]["is_admin"] is True
    assert by_email["bob.brown@example.com"]["is_billing_admin"] is True
    assert by_email["carol.chen@example.com"]["is_admin"] is False
    assert by_email["carol.chen@example.com"]["is_billing_admin"] is False


def test_fetch_all_users_handles_missing_names(monkeypatch):
    monkeypatch.setattr(export_mod.requests, "get", _fake_get)
    users = export_mod.fetch_all_users("https://api.solace.cloud", "fake-token")
    carol = next(u for u in users if u["email"] == "carol.chen@example.com")
    assert carol["first_name"] == ""
    assert carol["last_name"] == ""


def test_fetch_all_users_role_separator_is_configurable(monkeypatch):
    monkeypatch.setattr(export_mod.requests, "get", _fake_get)
    users = export_mod.fetch_all_users("https://api.solace.cloud", "fake-token", role_sep=", ")
    bob = next(u for u in users if u["email"] == "bob.brown@example.com")
    assert bob["roles"] == "billing-administrator, event-portal-user"
    assert bob["role_count"] == 2


def test_fetch_all_users_sap_admin_role_counts_as_admin(monkeypatch):
    """Regression test: is_admin must recognize SAP AEM's top-level admin
    role (sap-organization-administrator), not just Solace Cloud's
    "administrator" — this was previously a known, documented gap."""
    sap_admin_user = {
        "id": "u9", "organizationId": "aem-org", "firstName": "Dana", "lastName": "SAP",
        "email": "dana.sap@example-sap.com", "roles": ["sap-organization-administrator"],
        "groups": [], "state": "ACTIVE",
    }

    def fake_get_single_page(url, headers=None, params=None, timeout=None):
        return _make_response([sap_admin_user], 1, 1, None)

    monkeypatch.setattr(export_mod.requests, "get", fake_get_single_page)
    users = export_mod.fetch_all_users("https://api.solace.cloud", "fake-token")
    assert users[0]["is_admin"] is True


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

@pytest.fixture
def exported_users(monkeypatch):
    monkeypatch.setattr(export_mod.requests, "get", _fake_get)
    return export_mod.fetch_all_users("https://api.solace.cloud", "fake-token")


def test_write_csv(tmp_path, exported_users):
    path = tmp_path / "out.csv"
    export_mod.write_csv(exported_users, str(path))
    with open(path, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 3
    assert rows[0]["email"] == "alice.anders@example.com"


def test_write_json(tmp_path, exported_users):
    path = tmp_path / "out.json"
    export_mod.write_json(exported_users, str(path))
    data = json.loads(path.read_text())
    assert {"exported_at", "total_users", "users"} <= data.keys()
    assert data["total_users"] == 3
    assert len(data["users"]) == 3


def test_write_excel(tmp_path, exported_users):
    path = tmp_path / "out.xlsx"
    export_mod.write_excel(exported_users, str(path))
    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == {"All Users", "Admins", "Role Summary"}
    # header row + 1 admin (alice) = 2 rows
    assert wb["Admins"].max_row == 2


def test_write_excel_role_summary_respects_custom_separator(tmp_path, monkeypatch):
    """Regression test: Role Summary used to hardcode ' | ' as the split
    separator regardless of --role-separator, silently breaking role
    counts for any other separator (including the README's own
    documented example, --role-separator ", ")."""
    monkeypatch.setattr(export_mod.requests, "get", _fake_get)
    users = export_mod.fetch_all_users("https://api.solace.cloud", "fake-token", role_sep=", ")

    path = tmp_path / "out.xlsx"
    export_mod.write_excel(users, str(path), role_sep=", ")
    wb = openpyxl.load_workbook(path)

    role_summary = {
        row[0]: row[1]
        for row in wb["Role Summary"].iter_rows(min_row=2, values_only=True)
    }
    # bob has 2 roles ("billing-administrator, event-portal-user"); with the
    # bug, the whole comma-joined string would be counted as one bogus
    # "role" instead of being split into these two.
    assert role_summary.get("billing-administrator") == 1
    assert role_summary.get("event-portal-user") >= 1
    assert "billing-administrator, event-portal-user" not in role_summary


# ---------------------------------------------------------------------------
# _get_with_retry() — retry/backoff on transient failures
# ---------------------------------------------------------------------------

def test_get_with_retry_succeeds_first_try(monkeypatch):
    monkeypatch.setattr(export_mod.requests, "get", _fake_get)
    resp = export_mod._get_with_retry(
        "https://api.solace.cloud/api/v2/platform/users",
        headers={}, params={"pageNumber": 1, "pageSize": 100},
    )
    assert resp.json()["data"] == PAGE_1_USERS


def test_get_with_retry_recovers_from_connection_error(monkeypatch):
    monkeypatch.setattr(export_mod.time, "sleep", lambda s: None)
    calls = {"n": 0}

    def flaky_get(url, headers=None, params=None, timeout=None):
        calls["n"] += 1
        if calls["n"] < 2:
            raise export_mod.requests.exceptions.ConnectionError("boom")
        return _make_response(PAGE_1_USERS, 1, 1, None)

    monkeypatch.setattr(export_mod.requests, "get", flaky_get)
    resp = export_mod._get_with_retry("https://api.solace.cloud/x", headers={}, params={})
    assert resp.json()["data"] == PAGE_1_USERS
    assert calls["n"] == 2


def test_get_with_retry_recovers_from_retryable_http_status(monkeypatch):
    monkeypatch.setattr(export_mod.time, "sleep", lambda s: None)
    calls = {"n": 0}

    def flaky_get(url, headers=None, params=None, timeout=None):
        calls["n"] += 1
        if calls["n"] < 2:
            resp = MagicMock()
            resp.status_code = 503
            resp.text = "Service Unavailable"
            err = export_mod.requests.exceptions.HTTPError(response=resp)
            raise err
        return _make_response(PAGE_1_USERS, 1, 1, None)

    monkeypatch.setattr(export_mod.requests, "get", flaky_get)
    resp = export_mod._get_with_retry("https://api.solace.cloud/x", headers={}, params={})
    assert resp.json()["data"] == PAGE_1_USERS
    assert calls["n"] == 2


def test_get_with_retry_does_not_retry_non_retryable_http_status(monkeypatch):
    """A 404 (or other non-retryable status) must fail immediately —
    retrying it would never succeed."""
    monkeypatch.setattr(export_mod.time, "sleep", lambda s: None)
    calls = {"n": 0}

    def always_404(url, headers=None, params=None, timeout=None):
        calls["n"] += 1
        resp = MagicMock()
        resp.status_code = 404
        resp.text = "Not Found"
        raise export_mod.requests.exceptions.HTTPError(response=resp)

    monkeypatch.setattr(export_mod.requests, "get", always_404)
    with pytest.raises(SystemExit):
        export_mod._get_with_retry("https://api.solace.cloud/x", headers={}, params={})
    assert calls["n"] == 1


def test_get_with_retry_exhausts_attempts_and_exits(monkeypatch):
    monkeypatch.setattr(export_mod.time, "sleep", lambda s: None)
    calls = {"n": 0}

    def always_timeout(url, headers=None, params=None, timeout=None):
        calls["n"] += 1
        raise export_mod.requests.exceptions.Timeout("too slow")

    monkeypatch.setattr(export_mod.requests, "get", always_timeout)
    with pytest.raises(SystemExit):
        export_mod._get_with_retry("https://api.solace.cloud/x", headers={}, params={})
    assert calls["n"] == export_mod.MAX_ATTEMPTS


# ---------------------------------------------------------------------------
# End-to-end: exporter + verify_export.py, mirroring the README's
# `solace_cloud_export_script.py && verify_export.py` chain
# ---------------------------------------------------------------------------

def test_full_export_and_verify_roundtrip(tmp_path, monkeypatch):
    monkeypatch.setattr(export_mod.requests, "get", _fake_get)
    monkeypatch.chdir(tmp_path)

    monkeypatch.setattr(sys, "argv", [
        "solace_cloud_export_script.py", "--token", "fake-token", "--output-dir", "./output",
    ])
    export_mod.main()

    run_dirs = list((tmp_path / "output").iterdir())
    assert len(run_dirs) == 1

    monkeypatch.setattr(sys, "argv", ["verify_export.py", "--output-dir", "./output"])
    with pytest.raises(SystemExit) as exc_info:
        verify_mod.main()
    assert exc_info.value.code == 0


def test_verify_export_fails_on_inconsistent_data(tmp_path, monkeypatch):
    """verify_export.py must exit 1 (not silently pass) when the formats disagree."""
    run_dir = tmp_path / "output" / "20260101000000"
    run_dir.mkdir(parents=True)
    (run_dir / verify_mod.CSV_NAME).write_text("user_id,email\nu1,a@example.com\n")
    (run_dir / verify_mod.JSON_NAME).write_text(json.dumps({
        "exported_at": "2026-01-01T00:00:00+00:00",
        "total_users": 2,  # deliberately wrong — should be 1
        "users": [{"user_id": "u1", "email": "a@example.com"}],
    }))

    monkeypatch.setattr(sys, "argv", ["verify_export.py", "--output-dir", str(tmp_path / "output")])
    with pytest.raises(SystemExit) as exc_info:
        verify_mod.main()
    assert exc_info.value.code == 1
